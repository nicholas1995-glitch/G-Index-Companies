# Il tuo codice originale inizia qui
import requests
from lxml import html
import random
import os
import json
import logging
import time
from google.cloud import firestore
from google.oauth2 import service_account
from datetime import datetime
import openpyxl
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders

# Configura il logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[logging.StreamHandler()]
)
logger = logging.getLogger(__name__)

# Percorso del Secret File montato da Render
FIRESTORE_KEY_PATH = '/etc/secrets/firestore-key.json'

# Inizializza Firestore con le credenziali
try:
    if not os.path.exists(FIRESTORE_KEY_PATH):
        logger.error(f"File delle credenziali di Firestore non trovato: {FIRESTORE_KEY_PATH}")
        exit(1)
    credentials = service_account.Credentials.from_service_account_file(FIRESTORE_KEY_PATH)
    db = firestore.Client(credentials=credentials)
    logger.info("Connessione a Firestore riuscita")
except Exception as e:
    logger.error(f"Errore nella connessione a Firestore: {e}")
    exit(1)

# Lista aziende con nome completo, ticker e ISIN
companies_info = {
   "DHL.DE": {"name": "Deutsche Post", "isin": "DE0005552004"},
    "CBK.DE": {"name": "Commerzbank", "isin": "DE000CBK1001"},
    "DB1.DE": {"name": "Deutsche Börse", "isin": "DE0005810055"},
    "BMW.DE": {"name": "BMW", "isin": "DE0005190003"},
    "DTE.DE": {"name": "Deutsche Telekom", "isin": "DE0005557508"},
    "MRK.DE": {"name": "Merck", "isin": "DE0006599905"},
    "IFX.DE": {"name": "Infineon", "isin": "DE0006231004"},
    "MUV2.DE": {"name": "Münchener Rück", "isin": "DE0008430026"},
    "SAP.DE": {"name": "SAP", "isin": "DE0007164600"},
    "RHM.DE": {"name": "Rheinmetall", "isin": "DE0007030009"},
    "BEI.DE": {"name": "Beiersdorf", "isin": "DE0005200000"},
    "FRE.DE": {"name": "Fresenius SE", "isin": "DE0005785604"},
    "HNR1.DE": {"name": "Hannover Rück", "isin": "DE0008402215"},
    "RWE.DE": {"name": "RWE", "isin": "DE0007037129"},
    "SY1.DE": {"name": "Symrise", "isin": "DE000SYM9999"},
    "HEI.DE": {"name": "Heidelberg Materials", "isin": "DE0006047004"},
    "QIA.DE": {"name": "Qiagen", "isin": "NL0012169213"},
    "EOAN.DE": {"name": "E.ON", "isin": "DE000ENAG999"},
    "SRT3.DE": {"name": "Sartorius", "isin": "DE0007165631"},
    "VOW3.DE": {"name": "Volkswagen", "isin": "DE0007664039"},
    "ZAL.DE": {"name": "Zalando", "isin": "DE000ZAL1111"},
    "SHL.DE": {"name": "Siemens Healthineers", "isin": "DE000SHL1006"},
    "P911.DE": {"name": "Dr. Ing. h.c. F. Porsche AG", "isin": "DE000PAG9113"},
    "AC.PA": {"name": "Accor", "isin": "FR0000120404"},
    "AI.PA": {"name": "Air Liquide", "isin": "FR0000120073"},
    "CAP.PA": {"name": "Capgemini", "isin": "FR0000125338"},
    "CS.PA": {"name": "Axa", "isin": "FR0000120628"},
    "BN.PA": {"name": "Danone", "isin": "FR0000120644"},
    "ENGI.PA": {"name": "Engie", "isin": "FR0010208488"},
    "BNP.PA": {"name": "BNP Paribas", "isin": "FR0000131104"},
    "EN.PA": {"name": "Bouygues", "isin": "FR0000120503"},
    "MC.PA": {"name": "Louis Vuitton (LVMH)", "isin": "FR0000121014"},
    "OR.PA": {"name": "L'Oréal", "isin": "FR0000120321"},
    "SU.PA": {"name": "Schneider Electric", "isin": "FR0000121972"},
    "SGO.PA": {"name": "Saint-Gobain", "isin": "FR0000125007"},
    "CA.PA": {"name": "Carrefour", "isin": "FR0000120172"},
    "EL.PA": {"name": "EssilorLuxottica", "isin": "FR0000121667"},
    "RI.PA": {"name": "Pernod Ricard", "isin": "FR0000120693"},
    "ORA.PA": {"name": "Orange", "isin": "FR0000133308"},
    "GLE.PA": {"name": "Société Générale", "isin": "FR0000130809"},
    "SAN.PA": {"name": "Sanofi", "isin": "FR0000120578"},
    "VIV.PA": {"name": "Vivendi", "isin": "FR0000127771"},
    "DG.PA": {"name": "Vinci", "isin": "FR0000125486"},
    "RNO.PA": {"name": "Renault", "isin": "FR0000131906"},
    "VIE.PA": {"name": "Veolia Environnement", "isin": "FR0000124141"},
    "KER.PA": {"name": "Kering", "isin": "FR0000121485"},
    "HO.PA": {"name": "Thales", "isin": "FR0000121329"},
    "DSY.PA": {"name": "Dassault Systèmes", "isin": "FR0000130650"},
    "ERF.PA": {"name": "Eurofins Scientific SE", "isin": "FR0000038259"},
    "PUB.PA": {"name": "Publicis Groupe", "isin": "FR0000130577"},
    "RMS.PA": {"name": "Hermès International", "isin": "FR0000052292"},
    "SAF.PA": {"name": "Safran", "isin": "FR0000073272"},
    "LR.PA": {"name": "Legrand", "isin": "FR0010307819"},
    "STLAM.MI": {"name": "Stellantis", "isin": "NL00150001Q9"},
    "G.MI": {"name": "Assicurazioni Generali", "isin": "IT0000062072"},
    "SRG.MI": {"name": "Snam Rete Gas", "isin": "IT0003153415"},
    "UCG.MI": {"name": "UniCredit", "isin": "IT0005239360"},
    "LDO.MI": {"name": "Leonardo", "isin": "IT0003856405"},
    "STMMI.MI": {"name": "STMicroelectronics", "isin": "NL0000226223"},
    "ENEL.MI": {"name": "Enel", "isin": "IT0003128367"},
    "ENI.MI": {"name": "Eni", "isin": "IT0003132476"},
    "TRN.MI": {"name": "Terna", "isin": "IT0003242622"},
    "SPM.MI": {"name": "Saipem", "isin": "IT0005252140"},
    "CPR.MI": {"name": "Campari", "isin": "NL0015435975"},
    "DIA.MI": {"name": "DiaSorin", "isin": "IT0003492391"},
    "REC.MI": {"name": "Recordati", "isin": "IT0003828271"},
    "MONC.MI": {"name": "Moncler", "isin": "IT0004965148"},
    "INW.MI": {"name": "Inwit", "isin": "IT0005090300"},
}

# Elenco User-Agent (NON modifico quelli già presenti)
USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/90.0.4430.93 Safari/537.36",
    "Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:88.0) Gecko/20100101 Firefox/88.0",
    "Mozilla/5.0 (iPhone; CPU iPhone OS 14_6 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/14.0 Mobile/15E148 Safari/604.1",
    # Gli altri User-Agent che avevi...
]

session = requests.Session()
session.headers.update({"User-Agent": random.choice(USER_AGENTS)})

# Funzione per scraping (NON modifico nulla qui)
def scrape_stock_data(ticker):
    try:
        logger.info(f"Inizio scraping per {ticker}")
        url_main = f"https://finance.yahoo.com/quote/{ticker}/"
        url_stats = f"https://finance.yahoo.com/quote/{ticker}/key-statistics/"

        session = requests.Session()
        session.headers.update({"User-Agent": random.choice(USER_AGENTS)})

        # Richiesta per la pagina principale
        logger.info(f"Richiesta GET a {url_main}")
        response_main = session.get(url_main, timeout=20)
        time.sleep(2)  # Attendi 2 secondi dopo la richiesta
        response_main.raise_for_status()
        tree_main = html.fromstring(response_main.content)
        logger.info(f"Risposta ricevuta da {url_main} (Status Code: {response_main.status_code})")

        # Richiesta per la pagina delle statistiche
        logger.info(f"Richiesta GET a {url_stats}")
        response_stats = session.get(url_stats, timeout=20)
        time.sleep(5) # Attendi 5 secondi dopo la richiesta
        response_stats.raise_for_status()
        tree_stats = html.fromstring(response_stats.content)
        logger.info(f"Risposta ricevuta da {url_stats} (Status Code: {response_stats.status_code})")

        # Usa i nuovi XPath
        pe_ratio = tree_main.xpath('//*[@id="nimbus-app"]/section/section/section/article/div[2]/ul/li[11]/span[2]/fin-streamer/text()')
        pb_ratio = tree_stats.xpath('//*[@id="nimbus-app"]/section/section/section/article/section[2]/div/table/tbody/tr[7]/td[2]/text()')
        peg_ratio = tree_stats.xpath('//*[@id="nimbus-app"]/section/section/section/article/section[2]/div/table/tbody/tr[5]/td[2]/text()')

        # Aggiungi log dei valori estratti
        logger.debug(f"P/E Ratio estratto: {pe_ratio}")
        logger.debug(f"P/Book Ratio estratto: {pb_ratio}")
        logger.debug(f"PEG Ratio estratto: {peg_ratio}")

        # Verifica se i valori sono stati estratti correttamente
        if not pe_ratio or not pb_ratio or not peg_ratio:
            logger.warning(f"Dati incompleti per {ticker}: PE={pe_ratio}, PB={pb_ratio}, PEG={peg_ratio}")

        logger.info(f"Scraping completato per {ticker}")

        return {
            "P/E Ratio": pe_ratio[0].strip() if pe_ratio else "--",
            "P/Book Ratio": pb_ratio[0].strip() if pb_ratio else "--",
            "PEG Ratio (5y)": peg_ratio[0].strip() if peg_ratio else "--",
            "Ticker": ticker,
            "Full Name": companies_info[ticker]["name"],
            "ISIN": companies_info[ticker]["isin"],
            "Calcolo": f"((P/E: {pe_ratio[0] if pe_ratio else '--'} / 15) * 0.4) + (P/B: {pb_ratio[0] if pb_ratio else '--'} * 0.4) + (PEG: {peg_ratio[0] if peg_ratio else '--'} * 0.2)",
        }
    except Exception as e:
        logger.error(f"Errore durante lo scraping di {ticker}: {e}")
        return {
            "P/E Ratio": "--",
            "P/Book Ratio": "--",
            "PEG Ratio (5y)": "--",
            "Ticker": ticker,
            "Full Name": companies_info[ticker]["name"],
            "ISIN": companies_info[ticker]["isin"],
            "Calcolo": "Errore nel calcolo",
        }

def calculate_g_index(company):
    try:
        pe = float(company["P/E Ratio"].replace(",", ".")) if company["P/E Ratio"] not in ("--", "") else None
        pb = float(company["P/Book Ratio"].replace(",", ".")) if company["P/Book Ratio"] not in ("--", "") else None
        peg = float(company["PEG Ratio (5y)"].replace(",", ".")) if company["PEG Ratio (5y)"] not in ("--", "") else None

        # Verifica se tutti i valori sono disponibili
        if pe is None or pb is None or peg is None:
            logger.warning(f"Dati incompleti per il calcolo dell'Indice G di {company['Ticker']}")
            return "--"

        g_index = round(((pe / 15) * 0.4) + (pb * 0.4) + (peg * 0.2), 2)
        return g_index
    except Exception as e:
        logger.error(f"Errore nel calcolo dell'Indice G per {company['Ticker']}: {e}")
        return "--"

# Nuove funzioni per gestire Excel e invio email (aggiunte dopo il codice originale)

def generate_excel(data):
    """
    Genera un file Excel con i dati aggiornati, includendo una nuova sheet per l'Indice G.
    """
    logger.info("Inizio generazione file Excel")
    file_name = "dati_aziende.xlsx"
    date = datetime.now().strftime("%d-%m-%Y")  # Cambia il separatore per evitare errori

    # Correggi i titoli dei fogli
    sheet_names = {
        "P/E": "P_E",
        "P/BOOK": "P_BOOK",
        "PEG RATIO 5Y": "PEG_RATIO_5Y",
        "Indice G": "Indice_G"  # Aggiungi la nuova sheet
    }

    # Crea il file Excel se non esiste
    if not os.path.exists(file_name):
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        for original_title in sheet_names:
            sheet = wb.create_sheet(sheet_names[original_title])
            if original_title != "Indice G":
                sheet.append(["Nome Azienda", "Ticker", date])
            else:
                sheet.append(["Nome Azienda", "Ticker", "Indice G", date])

        wb.save(file_name)

    # Carica il workbook esistente
    wb = openpyxl.load_workbook(file_name)

    # Aggiorna le sheet P/E, P/Book e PEG Ratio
    for original_title, metric in zip(["P/E", "P/BOOK", "PEG RATIO 5Y"], ["P/E Ratio", "P/Book Ratio", "PEG Ratio (5y)"]):
        sheet = wb[sheet_names[original_title]]
        col = sheet.max_column + 1
        sheet.cell(1, col, date)

        for idx, company in enumerate(data, start=2):
            sheet.cell(idx, 1, company["Full Name"])
            sheet.cell(idx, 2, company["Ticker"])
            value = company[metric]
            if value != "--":
                try:
                    value = float(value.replace(",", "."))  # Converti in float
                except:
                    pass
            cell = sheet.cell(idx, col, value)
            if isinstance(value, float):
                cell.number_format = '0.00'  # Imposta il formato con due decimali

    # Aggiungi o aggiorna la sheet Indice_G
    sheet_g = wb[sheet_names["Indice G"]] if sheet_names["Indice G"] in wb.sheetnames else wb.create_sheet(sheet_names["Indice G"])
    if sheet_g.max_row == 1 and sheet_g.max_column == 1:
        # Scrivi l'intestazione solo se la sheet è nuova
        sheet_g.append(["Nome Azienda", "Ticker", "Indice G", date])

    for company in data:
        nome = company["Full Name"]
        ticker = company["Ticker"]
        indice_g = company["Indice G"]

        # Cerca se la riga esiste già
        row = None
        for r in range(2, sheet_g.max_row + 1):
            if sheet_g.cell(r, 2).value == ticker:
                row = r
                break

        if row:
            # Aggiorna la riga esistente
            sheet_g.cell(row, 3, indice_g)
            sheet_g.cell(row, 4, date)
        else:
            # Aggiungi una nuova riga
            sheet_g.append([nome, ticker, indice_g, date])

    wb.save(file_name)
    logger.info("File Excel generato con successo")
    return file_name


def send_email(file_name):
    """
    Invia il file Excel via email.
    """
    logger.info("Inizio invio email")
    
    # Percorso del file segreto con la password
    EMAIL_PASSWORD_PATH = '/etc/secrets/EMAIL_PASSWORD'
    
    # Leggi la password dal file
    try:
        with open(EMAIL_PASSWORD_PATH, 'r') as f:
            password = f.read().strip()  # Usa .strip() per rimuovere eventuali spazi extra
    except Exception as e:
        logger.error(f"Errore nel caricamento della password email: {e}")
        exit(1)  # Termina l'esecuzione in caso di errore
    
    sender_email = "nicholas.gazzola@gmail.com"
    receiver_email = "nicholas.gazzola@gmail.com"
    
    subject = "Dati aggiornati aziende"
    body = "In allegato trovi i dati aggiornati delle aziende."

    msg = MIMEMultipart()
    msg['From'] = sender_email
    msg['To'] = receiver_email
    msg['Subject'] = subject
    msg.attach(MIMEText(body, 'plain'))

    # Aggiungi l'allegato
    with open(file_name, "rb") as attachment:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(attachment.read())
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", f"attachment; filename={file_name}")
    msg.attach(part)

    # Invio email
    try:
        with smtplib.SMTP('smtp.gmail.com', 587) as server:
            server.starttls()
            server.login(sender_email, password)
            server.sendmail(sender_email, receiver_email, msg.as_string())
            logger.info("Email inviata con successo")
    except Exception as e:
        logger.error(f"Errore nell'invio dell'email: {e}")

# Nuove funzioni aggiunte per gestire categorie e notifiche (Aggiunte qui)

def get_previous_ticker_categories():
    """
    Recupera le liste precedenti dei ticker per tutte e tre le categorie da Firestore.
    """
    try:
        doc_ref = db.collection("ticker_categories")
        docs = doc_ref.stream()
        previous = {}
        for doc in docs:
            previous[doc.id] = doc.to_dict()
        if not previous:
            logger.info("Nessun dato precedente trovato per ticker_categories.")
        return previous
    except Exception as e:
        logger.error(f"Errore nel recuperare ticker_categories da Firestore: {e}")
        return {}

def update_ticker_categories(new_green, new_orange, new_red):
    """
    Aggiorna le liste dei ticker per tutte e tre le categorie in Firestore.
    """
    try:
        doc_ref_green = db.collection("ticker_categories").document("green")
        doc_ref_orange = db.collection("ticker_categories").document("orange")
        doc_ref_red = db.collection("ticker_categories").document("red")
        
        green_dict = {company["Ticker"]: company["Indice G"] for company in new_green}
        orange_dict = {company["Ticker"]: company["Indice G"] for company in new_orange}
        red_dict = {company["Ticker"]: company["Indice G"] for company in new_red}
        
        doc_ref_green.set(green_dict)
        doc_ref_orange.set(orange_dict)
        doc_ref_red.set(red_dict)
        
        logger.info("Liste ticker_categories aggiornate in Firestore.")
    except Exception as e:
        logger.error(f"Errore nell'aggiornare ticker_categories in Firestore: {e}")

def compare_ticker_categories(previous, current):
    """
    Confronta le liste precedente e corrente dei ticker per tutte e tre le categorie.
    Ritorna un dizionario con entrate, uscite e trasferimenti.
    """
    categories = ['green', 'orange', 'red']
    changes = {
        'entered': [],
        'exited': [],
        'transferred': []
    }
    
    # Creiamo un dizionario che mappa ticker a categorie precedenti
    ticker_to_previous_category = {}
    for category in categories:
        if category in previous:
            for ticker in previous[category]:
                ticker_to_previous_category[ticker] = category
    
    # Creiamo un dizionario che mappa ticker a categorie correnti
    ticker_to_current_category = {}
    for category in categories:
        if category in current:
            for ticker in current[category]:
                ticker_to_current_category[ticker] = category
    
    # Ticker attuali e precedenti
    current_tickers = set(ticker_to_current_category.keys())
    previous_tickers = set(ticker_to_previous_category.keys())
    
    # Identificare i ticker entrati (presenti ora ma non prima)
    entered_tickers = current_tickers - previous_tickers
    for ticker in entered_tickers:
        category = ticker_to_current_category[ticker]
        indice_g = current[category][ticker]
        changes['entered'].append({
            'Ticker': ticker,
            'Categoria': category,
            'Indice G': indice_g
        })
    
    # Identificare i ticker usciti (presenti prima ma non ora)
    exited_tickers = previous_tickers - current_tickers
    for ticker in exited_tickers:
        category = ticker_to_previous_category[ticker]
        indice_g = previous[category][ticker]
        changes['exited'].append({
            'Ticker': ticker,
            'Categoria': category,
            'Indice G': indice_g
        })
    
    # Identificare i ticker trasferiti (presente in entrambe ma categoria diversa)
    common_tickers = current_tickers & previous_tickers
    for ticker in common_tickers:
        previous_category = ticker_to_previous_category[ticker]
        current_category = ticker_to_current_category[ticker]
        if previous_category != current_category:
            changes['transferred'].append({
                'Ticker': ticker,
                'Da': previous_category,
                'A': current_category,
                'Indice G': current[current_category][ticker]  # Corretto da 'category' a 'current_category'
            })
    
    return changes

def send_change_notification(changes):
    """
    Invia un'email notificando le modifiche nelle liste delle categorie.
    """
    if not changes['entered'] and not changes['exited'] and not changes['transferred']:
        logger.info("Nessuna modifica nelle liste delle categorie.")
        return

    logger.info("Invio notifica via email per modifiche nelle liste delle categorie.")

    # Percorso del file segreto con la password
    EMAIL_PASSWORD_PATH = '/etc/secrets/EMAIL_PASSWORD'

    # Leggi la password dal file
    try:
        with open(EMAIL_PASSWORD_PATH, 'r') as f:
            password = f.read().strip()
    except Exception as e:
        logger.error(f"Errore nel caricamento della password email: {e}")
        return

    sender_email = "nicholas.gazzola@gmail.com"
    receiver_email = "nicholas.gazzola@gmail.com"

    subject = "Aggiornamenti nelle Liste delle Aziende"

    body = "Ci sono stati aggiornamenti nelle liste delle aziende:\n\n"

    if changes['entered']:
        body += "**Aziende Entrate nelle Liste:**\n"
        for company in changes['entered']:
            body += f"- {company['Ticker']} nella categoria **{company['Categoria'].capitalize()}** con Indice G: {company['Indice G']}\n"
        body += "\n"

    if changes['exited']:
        body += "**Aziende Uscite dalle Liste:**\n"
        for company in changes['exited']:
            body += f"- {company['Ticker']} dalla categoria **{company['Categoria'].capitalize()}** con Indice G: {company['Indice G']}\n"
        body += "\n"

    if changes['transferred']:
        body += "**Aziende Trasferite tra le Liste:**\n"
        for company in changes['transferred']:
            body += f"- {company['Ticker']} da **{company['Da'].capitalize()}** a **{company['A'].capitalize()}** con Indice G: {company['Indice G']}\n"
        body += "\n"

    msg = MIMEMultipart()
    msg['From'] = sender_email
    msg['To'] = receiver_email
    msg['Subject'] = subject
    msg.attach(MIMEText(body, 'plain'))

    # Invio email
    try:
        with smtplib.SMTP('smtp.gmail.com', 587) as server:
            server.starttls()
            server.login(sender_email, password)
            server.sendmail(sender_email, receiver_email, msg.as_string())
            logger.info("Email di notifica inviata con successo.")
    except Exception as e:
        logger.error(f"Errore nell'invio dell'email di notifica: {e}")

# Funzioni esistenti rimangono inalterate

def update_data():
    """
    Esegue lo scraping dei dati, calcola l'indice G, salva i dati su Firestore,
    genera un Excel e invia il file via email.
    """
    logger.info("Inizio aggiornamento dei dati")
    tickers = list(companies_info.keys())
    companies = []

    for i in range(0, len(tickers), 5):  # Passo di 5
        batch = tickers[i:i + 5]
        logger.info(f"Elaborazione batch {i//5 + 1}: {batch}")
        for ticker in batch:
            logger.info(f"Elaborazione ticker: {ticker}")
            company_data = scrape_stock_data(ticker)
            company_data["Indice G"] = calculate_g_index(company_data)
            companies.append(company_data)
            logger.info(f"Ticker {ticker} elaborato con Indice G: {company_data['Indice G']}")
        if i + 5 < len(tickers):  # Attendi solo se ci sono altri batch
            logger.info(f"Attesa di 60 secondi prima di elaborare il prossimo batch...")
            time.sleep(60)

    green = [c for c in companies if c["Indice G"] != "--" and c["Indice G"] < 1]
    orange = [c for c in companies if c["Indice G"] != "--" and 1 <= c["Indice G"] < 1.5]
    red = [c for c in companies if c["Indice G"] != "--" and c["Indice G"] >= 1.5]

    green.sort(key=lambda x: x["Indice G"])
    orange.sort(key=lambda x: x["Indice G"])
    red.sort(key=lambda x: x["Indice G"])

    stored_data = {"green": green, "orange": orange, "red": red}

    # Nuove aggiunte: Gestione delle categorie e notifiche
    # Recupera le liste precedenti delle categorie
    previous_categories = get_previous_ticker_categories()

    # Costruisci il dizionario delle categorie correnti
    current_categories = {
        "green": {company["Ticker"]: company["Indice G"] for company in green},
        "orange": {company["Ticker"]: company["Indice G"] for company in orange},
        "red": {company["Ticker"]: company["Indice G"] for company in red}
    }

    # Confronta le categorie
    changes = compare_ticker_categories(previous_categories, current_categories)

    # Aggiorna le categorie in Firestore
    update_ticker_categories(green, orange, red)

    # Invia notifiche via email se ci sono cambiamenti
    send_change_notification(changes)

    # Salva i dati su Firestore
    try:
        db.collection("stored_data").document("data").set(stored_data)
        logger.info("Dati salvati su Firestore con successo")
    except Exception as e:
        logger.error(f"Errore durante il salvataggio dei dati su Firestore: {e}")

    # Genera l'Excel e invia via email
    file_name = generate_excel(companies)
    send_email(file_name)

    logger.info("Aggiornamento dei dati completato")

if __name__ == "__main__":
    update_data()
